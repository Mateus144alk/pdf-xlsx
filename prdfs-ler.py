import pdfplumber
import pandas as pd
import os
import re
import glob
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from txt_generator import gerar_arquivo_txt

def extrair_dados_pdf(path_pdf):
    dados = []
    tipo_rubrica = ""

    with pdfplumber.open(path_pdf) as pdf:
        for page in pdf.pages:
            texto = page.extract_text()
            if not texto:
                continue

            if "RUBRICA:" in texto:
                for linha in texto.splitlines():
                    if linha.strip().startswith("RUBRICA:"):
                        tipo_rubrica = linha.split("RUBRICA:")[1].strip()
                        break

            for linha in texto.splitlines():
                match = re.search(r"(.+?)\s+EST\d{2}\s+(\d{6,7})\s+R\s+0\s+([\d.,]+)\s+N", linha)
                if match:
                    nome = match.group(1).strip()
                    matricula = match.group(2)
                    valor = float(match.group(3).replace(".", "").replace(",", "."))
                    dados.append((nome, matricula, tipo_rubrica, valor))
    return dados

def consolidar_dados(lista_arquivos_pdf, caminho_saida, mes_folha):
    registros = []
    for arquivo in lista_arquivos_pdf:
        registros.extend(extrair_dados_pdf(arquivo))

    tipos_adicionais = sorted(set(r for _, _, r, _ in registros if "BASICO" not in r.upper()))

    consolidado = {}
    for nome, matricula, rubrica, valor in registros:
        if matricula not in consolidado:
            consolidado[matricula] = {
                "Nome": nome,
                "Matrícula": matricula,
                "Salário Básico": 0,
                "Sequência": mes_folha,  
                **{tipo: 0 for tipo in tipos_adicionais}
            }
        if "BASICO" in rubrica.upper():
            consolidado[matricula]["Salário Básico"] = valor
        else:
            consolidado[matricula][rubrica] = valor

    df = pd.DataFrame(consolidado.values())
    df.rename(columns={
        col: re.sub(r"^\d{5} - ", "", col).replace(" - LEI 11.091/05 AT", "")
        for col in df.columns if re.match(r"^\d{5} - ", col)
    }, inplace=True)

    df.to_excel(caminho_saida, index=False)
    messagebox.showinfo("Sucesso", f"Planilha criada com sucesso em:\n{caminho_saida}")
dados_retroativo = pd.DataFrame()

def selecionar_planilha_retroativa():
    global dados_retroativo
    caminho = filedialog.askopenfilename(filetypes=[("Planilhas Excel", "*.xlsx *.xls")])
    if caminho:
        entrada_planilha_retroativo.delete(0, tk.END)
        entrada_planilha_retroativo.insert(0, caminho)
        try:
            dados_retroativo = pd.read_excel(caminho)
            messagebox.showinfo("Sucesso", "Planilha de retroativos carregada com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar planilha:\n{e}")
    
def gerar_valores_retroativos():
    try:
        if dados_retroativo.empty:
            messagebox.showwarning("Aviso", "Nenhuma planilha de retroativo carregada.")
            return

        caminho_anterior = entrada_anterior.get()
        caminho_atual = entrada_atual.get()

        if not caminho_anterior or not caminho_atual:
            messagebox.showwarning("Aviso", "Selecione planilhas de meses para comparação (anterior e atual).")
            return

        df_ant = pd.read_excel(caminho_anterior)
        df_atu = pd.read_excel(caminho_atual)

        col_ref = [col for col in df_ant.columns if col not in ("Nome", "Matrícula")]
        if len(col_ref) != 1:
            messagebox.showerror("Erro", f"Apenas uma coluna de valor esperada. Encontrado: {col_ref}")
            return

        coluna_valor = col_ref[0]

        df_ant = df_ant[["Matrícula", coluna_valor]].rename(columns={coluna_valor: "Valor Antigo"})
        df_atu = df_atu[["Matrícula", coluna_valor]].rename(columns={coluna_valor: "Valor Atual"})

        df_diff = pd.merge(df_atu, df_ant, on="Matrícula")
        df_diff["x3"] = df_diff["Valor Atual"] - df_diff["Valor Antigo"]

        df_retro = pd.merge(dados_retroativo, df_diff, left_on="SIAPE", right_on="Matrícula", how="inner")
        df_retro["Valor Retroativo"] = (df_retro["x3"] * df_retro["MESES RETROATIVOS"]) + \
                                       ((df_retro["x3"] / 30) * df_retro["DIAS RETROATIVOS"])

        salvar_em = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if salvar_em:
            df_retro[["SIAPE", "NOME", "MESES RETROATIVOS", "DIAS RETROATIVOS", "Valor Retroativo"]].to_excel(salvar_em, index=False)
            messagebox.showinfo("Sucesso", f"Retroativo gerado com sucesso em:\n{salvar_em}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao gerar valores retroativos:\n{e}")


def comparar_planilhas(caminho_anterior, caminho_atual, caminho_saida, apenas_diferencas=False):
    df_abr = pd.read_excel(caminho_anterior)
    df_mai = pd.read_excel(caminho_atual)

    colunas_excluir = ["Nome", "Matrícula"]
    colunas_valores = [col for col in df_abr.columns if col not in colunas_excluir]

    for col in colunas_valores:
        df_abr[col] = pd.to_numeric(df_abr[col], errors="coerce").fillna(0)
        df_mai[col] = pd.to_numeric(df_mai[col], errors="coerce").fillna(0)

    df_comparado = pd.merge(
    df_abr, df_mai, on="Matrícula", suffixes=(" (Abr)", " (Mai)"), how="outer", indicator=True
   )
    df_comparado["Nome"] = df_comparado["Nome (Mai)"].combine_first(df_comparado["Nome (Abr)"])
    for col in colunas_valores:
        df_comparado[f"{col} (Diferença)"] = df_comparado[f"{col} (Mai)"] - df_comparado[f"{col} (Abr)"]

    colunas_ordenadas = ["Matrícula", "Nome"]
    for col in colunas_valores:
        colunas_ordenadas.extend([f"{col} (Abr)", f"{col} (Mai)", f"{col} (Diferença)"])

    df_final = df_comparado[colunas_ordenadas]

    if apenas_diferencas:
     diff_cols = [col for col in df_final.columns if col.endswith("(Diferença)")]
     df_final[diff_cols] = df_final[diff_cols].round(2)
     df_final = df_final[df_final[diff_cols].abs().sum(axis=1) > 0]



    df_final.to_excel(caminho_saida, index=False)
    messagebox.showinfo("Sucesso", f"Comparativo gerado com sucesso em:\n{caminho_saida}")
    wb = load_workbook(caminho_saida)
    ws = wb.active
    fill = PatternFill(start_color="ee0000", end_color="ee0000", fill_type="solid")  # vermelho claro

    for col in range(1, ws.max_column + 1):
        col_name = ws.cell(row=1, column=col).value
        if col_name and "(Diferença)" in col_name:
            for row in range(2, ws.max_row + 1):
                valor = ws.cell(row=row, column=col).value
                if isinstance(valor, (int, float)) and abs(valor) > 0.001:
                    ws.cell(row=row, column=col).fill = fill

    wb.save(caminho_saida)
janela = tk.Tk()
janela.title("Consolidador e Comparador de Salários")
janela.geometry("750x350")

def escolher_pasta_pdf():
    pasta = filedialog.askdirectory()
    if pasta:
        entrada_pasta.delete(0, tk.END)
        entrada_pasta.insert(0, pasta)

def salvar_excel_saida():
    if check_destino_pasta.get():
        # Selecionar apenas a pasta
        pasta = filedialog.askdirectory()
        if pasta:
            # Tenta pegar o número do mês digitado
            mes = entrada_mes.get().strip()
            if not mes.isdigit():
                mes = "XX"
            nome_arquivo = f"consolidado_mes_{mes}.xlsx"
            caminho = os.path.join(pasta, nome_arquivo)

            entrada_saida.delete(0, tk.END)
            entrada_saida.insert(0, caminho)
    else:
        # Permite salvar como arquivo manualmente
        caminho = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if caminho:
            entrada_saida.delete(0, tk.END)
            entrada_saida.insert(0, caminho)


    
def acao_consolidar():
    caminho = entrada_pasta.get().strip()
    saida = entrada_saida.get().strip()
    mes = entrada_mes.get().strip()

    if not caminho or not saida or not mes:
        messagebox.showwarning("Atenção", "Preencha todos os campos, incluindo o mês.")
        return

    if not mes.isdigit():
        messagebox.showerror("Erro", "O mês deve ser um número inteiro.")
        return

    mes_folha = int(mes)

    if check_arquivo_unico.get():
        if not os.path.isfile(caminho) or not caminho.lower().endswith(".pdf"):
            messagebox.showerror("Erro", "Selecione um arquivo PDF válido.")
            return
        arquivos = [caminho]
    else:
        if not os.path.isdir(caminho):
            messagebox.showerror("Erro", "Selecione uma pasta válida.")
            return
        arquivos = glob.glob(os.path.join(caminho, "*.pdf"))
        if not arquivos:
            messagebox.showwarning("Aviso", "Nenhum PDF encontrado na pasta.")
            return

    consolidar_dados(arquivos, saida, mes_folha)



def selecionar_planilha_siape():
    global dados_siape
    caminho = filedialog.askopenfilename(filetypes=[("Planilhas Excel", "*.xlsx *.xls")])
    if caminho:
        entrada_siape.delete(0, tk.END)
        entrada_siape.insert(0, caminho)
        try:
            dados_siape = pd.read_excel(caminho, usecols="A:H")
            dados_siape.columns = [
                "SIAPE", "MATRÍCULA ORIGEM", "NOME SERVIDOR",
                "SITUAÇÃO", "CARGO", "CLASSE", "PADRÃO", "DV SIAPE"
            ]
            messagebox.showinfo("Sucesso", "Planilha SIAPE carregada com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler a planilha SIAPE:\n{e}")

def selecionar_planilha_anterior():
    caminho = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
    if caminho:
        entrada_anterior.delete(0, tk.END)
        entrada_anterior.insert(0, caminho)

def selecionar_planilha_atual():
    caminho = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
    if caminho:
        entrada_atual.delete(0, tk.END)
        entrada_atual.insert(0, caminho)
def selecionar_consolidado_unico():
    if check_destino_pasta.get():
        # Modo salvar em uma pasta (gera o nome automaticamente)
        pasta = filedialog.askdirectory()
        if pasta:
            mes = entrada_mes.get().strip()
            if not mes.isdigit():
                mes = "XX"
            nome_arquivo = f"consolidado_mes_{mes}.xlsx"
            caminho = os.path.join(pasta, nome_arquivo)
            entrada_saida.delete(0, tk.END)
            entrada_saida.insert(0, caminho)
    else:
        # Modo selecionar um arquivo .xlsx normalmente
        caminho = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if caminho:
            entrada_saida.delete(0, tk.END)
            entrada_saida.insert(0, caminho)

def escolher_pdf_ou_pasta():
    if check_arquivo_unico.get():
        caminho = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if caminho:
            entrada_pasta.delete(0, tk.END)
            entrada_pasta.insert(0, caminho)
    else:
        pasta = filedialog.askdirectory()
        if pasta:
            entrada_pasta.delete(0, tk.END)
            entrada_pasta.insert(0, pasta)

def salvar_comparativo():
    caminho = filedialog.asksaveasfilename(defaultextension=".xlsx")
    if caminho:
        entrada_comparativo.delete(0, tk.END)
        entrada_comparativo.insert(0, caminho)
def gerar_carga_batch():
    global dados_siape
    caminho_consolidado = entrada_saida.get()
    rubrica = entrada_rubrica.get().strip()
    sequencia = entrada_seq_batch.get().strip()

    if not caminho_consolidado or dados_siape.empty or not rubrica or not sequencia:
        messagebox.showwarning("Atenção", "Preencha rubrica, sequência e carregue a planilha SIAPE e consolidada.")
        return

    if not rubrica.isdigit() or not sequencia.isdigit():
        messagebox.showerror("Erro", "Rubrica e Sequência devem conter apenas números.")
        return

    try:
        df_consolidado = pd.read_excel(caminho_consolidado)
        df_siape = dados_siape.copy()

        df_merge = pd.merge(
            df_consolidado,
            df_siape,
            left_on="Matrícula",
            right_on="SIAPE",
            how="inner"
        )

        if df_merge.empty:
            messagebox.showerror("Erro", "Nenhuma correspondência encontrada entre consolidação e SIAPE.")
            return

        registros_batch = []
        colunas_padrao = ["Nome", "Matrícula", "Sequência"]
        colunas_valor = [
            col for col in df_consolidado.columns
            if col not in colunas_padrao and not df_consolidado[col].fillna(0).eq(0).all()
        ]

        if len(colunas_valor) != 1:
            messagebox.showerror("Erro", f"A planilha consolidada deve conter apenas UMA coluna de valor (além de Nome/Matrícula). Encontrado: {colunas_valor}")
            return

        coluna_valor = colunas_valor[0]

        for _, row in df_merge.iterrows():
            valor = row.get(coluna_valor, 0)
            if pd.isna(valor) or valor == 0:
                continue

            registros_batch.append({
                "MatSiape": row["SIAPE"],
                "DVSiape": row["DV SIAPE"],
                "Comando": 4,
                "RendimentoDesconto": 1,
                "Rubrica": rubrica.zfill(5),
                "Sequencia": int(sequencia),
                "Valor": round(valor, 2),
                "MatriculaOrigem": row["MATRÍCULA ORIGEM"]
            })

        if not registros_batch:
            messagebox.showinfo("Aviso", "Nenhum valor válido para gerar a carga.")
            return

        df_batch = pd.DataFrame(registros_batch)

        # Geração do CSV
        caminho_csv = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if caminho_csv:
            df_batch.to_csv(caminho_csv, index=False, sep=";", decimal=",")
            messagebox.showinfo("Sucesso", f"Carga CSV gerada com sucesso em:\n{caminho_csv}")

            # Pergunta se deseja gerar o .txt
            if messagebox.askyesno("TXT", "Deseja também gerar o arquivo .txt Movi-Financ?"):
                config = {
                    "matriz_padrao": "00000",
                    "mes_pagto": "02",
                    "ano_pagto": "2025",
                    "mes_rubrica": "01",
                    "ano_rubrica": "2025",
                    "nome_instituicao": "MFINST",
                    "rubrica_arquivo": rubrica.zfill(5)
                }

                caminho_txt = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Arquivos TXT", "*.txt")])
                if caminho_txt:
                    try:
                        gerar_arquivo_txt(df_batch, caminho_txt, config)
                        messagebox.showinfo("Sucesso", f"Arquivo TXT gerado com sucesso em:\n{caminho_txt}")
                    except Exception as e:
                        messagebox.showerror("Erro", f"Erro ao gerar TXT:\n{e}")

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao gerar carga:\n{e}")


def acao_comparar():
    abr = entrada_anterior.get()
    mai = entrada_atual.get()
    saida = entrada_comparativo.get()
    somente_diff = var_diferencas.get()
    if not abr or not mai or not saida:
        messagebox.showwarning("Atenção", "Preencha todos os campos da comparação.")
        return
    comparar_planilhas(abr, mai, saida, apenas_diferencas=somente_diff)

# Seção: Consolidador
frame1 = tk.LabelFrame(janela, text="Agrupar dados pdf em planilha")
frame1.pack(fill="x", padx=10, pady=5)

entrada_pasta = tk.Entry(frame1, width=60)
entrada_pasta.pack(side="left", padx=5, pady=5)

check_arquivo_unico = tk.BooleanVar()
tk.Checkbutton(frame1, text="Selecionar apenas um PDF", variable=check_arquivo_unico).pack(side="left", padx=5)

tk.Button(frame1, text="Selecionar", command=lambda: escolher_pdf_ou_pasta()).pack(side="left", padx=5)
frame_retroativo = tk.LabelFrame(janela, text="Planilha de Dados Retroativos")
frame_retroativo.pack(fill="x", padx=10, pady=5)

entrada_planilha_retroativo = tk.Entry(frame_retroativo, width=60)
entrada_planilha_retroativo.pack(side="left", padx=5)

tk.Button(frame_retroativo, text="Selecionar Planilha Retroativo", command=lambda: selecionar_planilha_retroativa()).pack(side="left")


entrada_saida = tk.Entry(janela, width=60)
entrada_saida.pack(padx=15)
check_destino_pasta = tk.BooleanVar()
tk.Checkbutton(janela, text="Salvar em uma pasta (gerar nome automático)", variable=check_destino_pasta).pack()
frame_mes = tk.Frame(janela)
frame_mes.pack(padx=10, pady=5)

tk.Label(frame_mes, text="Mês da folha (número):").pack(side="left")
entrada_mes = tk.Entry(frame_mes, width=10)
entrada_mes.pack(side="left", padx=5)

tk.Button(janela, text="Selecionar Pasta Arquivo Consolidado (Excel)", command=lambda: selecionar_consolidado_unico()).pack()
tk.Button(janela, text="📥 Gerar Planilha", bg="#4CAF50", fg="white", command=acao_consolidar).pack(pady=10)

frame3 = tk.LabelFrame(janela, text="Importar dados extraídos do SIAPE")
frame3.pack(fill="x", padx=10, pady=5)

entrada_siape = tk.Entry(frame3, width=60)
entrada_siape.pack(side="left", padx=5)
tk.Button(frame3, text="Selecionar Planilha SIAPE", command=lambda: selecionar_planilha_siape()).pack(side="left")
frame4 = tk.LabelFrame(janela, text="Gerar Carga Batch")
frame4.pack(fill="x", padx=10, pady=5)

tk.Label(frame4, text="Rubrica:").pack(side="left", padx=5)
entrada_rubrica = tk.Entry(frame4, width=10)
entrada_rubrica.pack(side="left")

tk.Label(frame4, text="Sequência:").pack(side="left", padx=5)
entrada_seq_batch = tk.Entry(frame4, width=5)
entrada_seq_batch.pack(side="left", padx=5)

tk.Button(frame4, text="💾 Gerar Carga CSV", bg="#FF9800", fg="white", command=lambda: gerar_carga_batch()).pack(side="left", padx=10)

# Seção: Comparar
frame2 = tk.LabelFrame(janela, text="Comparar Planilhas (anterior vs. atual)")
frame2.pack(fill="x", padx=10, pady=5)

entrada_anterior = tk.Entry(frame2, width=60)
entrada_anterior.pack(side="left", padx=5)
tk.Button(frame2, text="Planilha Anterior", command=selecionar_planilha_anterior).pack(side="left")

entrada_atual = tk.Entry(janela, width=60)
entrada_atual.pack(padx=15)
tk.Button(janela, text="Planilha Atual", command=selecionar_planilha_atual).pack()

entrada_comparativo = tk.Entry(janela, width=60)
entrada_comparativo.pack(padx=15)
tk.Button(janela, text="Salvar Comparativo", command=salvar_comparativo).pack()

var_diferencas = tk.BooleanVar()
tk.Checkbutton(janela, text="Exibir apenas diferenças", variable=var_diferencas).pack()
tk.Button(janela, text="📊 Comparar Meses", bg="#2196F3", fg="white", command=acao_comparar).pack(pady=10)

janela.mainloop()
