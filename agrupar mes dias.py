import pandas as pd
from openpyxl import load_workbook

# Caminhos
arquivo_prog = "C:/Ler arquivo/pdf-xlsx/3 - Pgto Retroativo PROGRESSÕES - Atualizado.xlsx"
arquivo_datas = "C:/Ler arquivo/pdf-xlsx/TODOS - PROGRESSÃO MÉRITO.ods"

# 1. Ler e preparar os dados
try:
    # Ler arquivo ODS
    df_datas = pd.read_excel(arquivo_datas, engine="odf")
    
    # Verificar colunas (convertendo nomes para maiúsculas para padronizar)
    df_datas.columns = df_datas.columns.str.upper()
    
    # Verificar se as colunas necessárias existem
    required_columns = {'MAT. SIAPE', 'DIA', 'MÊS'}
    if not required_columns.issubset(df_datas.columns):
        missing = required_columns - set(df_datas.columns)
        raise ValueError(f"Colunas faltando: {missing}")

    # Processar dados
    df_datas['MAT. SIAPE'] = df_datas['MAT. SIAPE'].astype(str).str.strip()
    df_datas['DIA'] = df_datas['DIA'].astype(str).str.strip()
    df_datas['MÊS'] = df_datas['MÊS'].astype(str).str.strip()

    # 2. Tratar valores duplicados (manter a primeira ocorrência)
    df_datas = df_datas.drop_duplicates(subset=['MAT. SIAPE'], keep='first')
    
    # 3. Carregar arquivo Excel de progressões
    wb = load_workbook(arquivo_prog)
    ws = wb.active
    
    # 4. Criar dicionário de mapeamento
    mapa_datas = df_datas.set_index('MAT. SIAPE')[['DIA', 'MÊS']].to_dict('index')
    
    # 5. Preencher dados na planilha de progressões
    for row in range(3, ws.max_row + 1):
        siape_cell = ws.cell(row=row, column=2).value  # Coluna B (2)
        if siape_cell:
            siape = str(siape_cell).strip()
            if siape in mapa_datas:
                dia = mapa_datas[siape]['DIA']
                mes = mapa_datas[siape]['MÊS']
                ws.cell(row=row, column=34, value=dia)  # Coluna AH
                ws.cell(row=row, column=35, value=mes)  # Coluna AI
    
    # 6. Salvar com backup
    backup_path = arquivo_prog.replace('.xlsx', '_BACKUP.xlsx')
    wb.save(backup_path)
    print(f"Backup criado em: {backup_path}")
    
    wb.save(arquivo_prog)
    print(f"Arquivo atualizado: {arquivo_prog}")
    print(f"Total de registros processados: {len(mapa_datas)}")

except Exception as e:
    print(f"\nErro durante a execução: {str(e)}")
    print("\nColunas disponíveis no arquivo ODS:")
    print(df_datas.columns.tolist() if 'df_datas' in locals() else "Arquivo não pôde ser lido")